import json, sys
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def main():
    json_path = sys.argv[1]
    with open(json_path) as f:
        squads = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)

    hf = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    hfont = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    tfont = Font(name="Calibri", bold=True, size=14, color="1F2937")
    sfont = Font(name="Calibri", italic=True, size=10, color="6B7280")
    df = Font(name="Calibri", size=10)
    pf = Font(name="Calibri", size=10, color="6B7280")
    bf = Font(name="Calibri", bold=True, size=10, color="374151")
    tb = Border(
        left=Side(style="thin", color="D1D5DB"), right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"), bottom=Side(style="thin", color="D1D5DB"),
    )
    ct = Alignment(horizontal="center", vertical="center")
    la = Alignment(horizontal="left", vertical="center")

    sc = {
        "Assault": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "Ambush": PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid"),
        "Medic": PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid"),
        "Kage": PatternFill(start_color="EDE9FE", end_color="EDE9FE", fill_type="solid"),
        "HQ": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
    }

    for squad in squads:
        ws = wb.create_sheet(title=squad["name"])
        players = squad["players"]
        n = len(players)
        sf = sc[squad["name"]]

        all_c = Counter(); t1c = Counter(); t2c = Counter(); t3c = Counter()
        for p in players:
            for k, c in [("talent1", t1c), ("talent2", t2c), ("talent3", t3c)]:
                v = p.get(k, "")
                if v and v != "None": c[v] += 1; all_c[v] += 1

        ws.merge_cells("A1:F1")
        ws["A1"].value = f"SHADOW WAR — {squad['name'].upper()} ({n} players)"
        ws["A1"].font = tfont; ws["A1"].alignment = ct
        ws.row_dimensions[1].height = 30
        ws.merge_cells("A2:F2")
        ws["A2"].value = f"Buff: {squad['buff']}  |  Debuff: {squad['debuff']}"
        ws["A2"].font = sfont; ws["A2"].alignment = Alignment(horizontal="center")

        row = 4
        ws.cell(row=row, column=1, value="Player List").font = bf
        row += 1
        for ci, h in enumerate(["Rank", "Player", "ID", "Talent 1", "Talent 2", "Talent 3"], 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = hfont; c.fill = hf; c.alignment = ct; c.border = tb

        for p in players:
            row += 1
            vals = [p["rank"], p["name"], p["id"],
                    "" if p.get("talent1") in (None, "", "None") else p["talent1"],
                    "" if p.get("talent2") in (None, "", "None") else p["talent2"],
                    "" if p.get("talent3") in (None, "", "None") else p["talent3"]]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=v)
                c.font = df; c.border = tb
                c.alignment = ct if ci in (1,3) else la
                if p["rank"] % 2 == 0: c.fill = sf

        row += 2
        ws.cell(row=row, column=1, value="Talent Count").font = bf
        for label, counter in [("All (T1+T2+T3)", all_c), ("Talent 1", t1c), ("Talent 2", t2c), ("Talent 3", t3c)]:
            row += 1
            ws.cell(row=row, column=1, value=label).font = bf
            row += 1
            for ci, h in enumerate(["Talent", "Count", "%"], 1):
                c = ws.cell(row=row, column=ci, value=h)
                c.font = hfont; c.fill = hf; c.alignment = ct; c.border = tb
            denom = n*3 if label.startswith("All") else n
            for talent, count in counter.most_common():
                row += 1
                ws.cell(row=row, column=1, value=talent).font = df
                ws.cell(row=row, column=1).border = tb
                ws.cell(row=row, column=2, value=count).font = df
                ws.cell(row=row, column=2).alignment = ct
                ws.cell(row=row, column=2).border = tb
                ws.cell(row=row, column=3, value=f"{count/denom*100:.1f}%").font = pf
                ws.cell(row=row, column=3).alignment = ct
                ws.cell(row=row, column=3).border = tb
            row += 1

        for col, w in {"A":28,"B":22,"C":10,"D":24,"E":24,"F":24}.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A6"

    import os
    outpath = os.path.expanduser("~/SW_Talent.xlsx")
    wb.save(outpath)
    print(outpath)

if __name__ == "__main__":
    main()
